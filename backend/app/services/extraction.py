"""Deterministic CSV/XLSX extraction and table detection."""

import csv
import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def detect_file_type(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    if suffix == ".pdf":
        return "pdf"
    return "unknown"


def extract_csv(data: bytes) -> dict[str, Any]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [list(row) for row in reader]
    return {
        "sheets": [
            {
                "name": "Sheet1",
                "rows": rows,
            }
        ]
    }


def extract_xlsx(data: bytes) -> dict[str, Any]:
    workbook = load_workbook(filename=io.BytesIO(data), data_only=True, read_only=True)
    sheets = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if cell is None else cell) for cell in row])
        sheets.append({"name": sheet_name, "rows": rows})
    return {"sheets": sheets}


def extract_workbook(filename: str, data: bytes) -> dict[str, Any]:
    file_type = detect_file_type(filename)
    if file_type == "csv":
        return extract_csv(data)
    if file_type == "xlsx":
        return extract_xlsx(data)
    raise ValueError(f"Unsupported file type for extraction: {file_type}")


def nonempty_row(row: list[Any]) -> bool:
    return any(str(cell).strip() != "" for cell in row)


def detect_header_row(rows: list[list[Any]], max_scan: int = 30) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:max_scan]):
        if not nonempty_row(row):
            continue
        text_cells = [str(c).strip() for c in row if str(c).strip()]
        if not text_cells:
            continue
        alphaish = sum(1 for c in text_cells if any(ch.isalpha() for ch in c))
        score = alphaish * 2 + len(text_cells)
        # Prefer rows that look like headers over numeric data rows.
        numeric = sum(1 for c in text_cells if _looks_numeric(c))
        score -= numeric
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def _looks_numeric(value: str) -> bool:
    cleaned = value.replace(",", "").replace(".", "").replace("-", "").replace("$", "")
    return cleaned.isdigit()


def build_structural_summary(
    workbook: dict[str, Any],
    *,
    default_currency: str,
    guessed_source_type: str = "bank_statement",
) -> dict[str, Any]:
    sheets = workbook["sheets"]
    selected = max(sheets, key=lambda s: sum(1 for r in s["rows"] if nonempty_row(r)))
    rows = selected["rows"]
    header_row = detect_header_row(rows)
    headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[header_row])]
    data_start = header_row + 1
    sample_rows = []
    for row in rows[data_start : data_start + 8]:
        if not nonempty_row(row):
            continue
        sample_rows.append(
            {
                headers[i] if i < len(headers) else f"col_{i}": _mask_sensitive(str(cell))
                for i, cell in enumerate(row)
                if i < len(headers) and str(cell).strip() != ""
            }
        )
    return {
        "selected_sheet": selected["name"],
        "sheet_names": [s["name"] for s in sheets],
        "header_row": header_row,
        "data_start_row": data_start,
        "headers": headers,
        "sample_rows": sample_rows,
        "default_currency": default_currency,
        "guessed_source_type": guessed_source_type,
        "warnings": [],
        "row_count_estimate": sum(1 for r in rows[data_start:] if nonempty_row(r)),
    }


def _mask_sensitive(value: str) -> str:
    import re

    return re.sub(r"(\d{4}[-\s]?){3}\d{4}", "****-****-****-****", value)
